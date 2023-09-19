import openpyxl
import pyautogui
import logging
import time

# Configurar o logger
logging.basicConfig(filename='log.txt', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Função para aguardar até que uma das duas imagens apareça na tela
def wait_for_images(images, timeout):
    start_time = pyautogui.time.time()
    while pyautogui.time.time() - start_time < timeout:
        for image in images:
            if pyautogui.locateOnScreen(image):
                return True
    return False

# Passo 1: Solicitar ao usuário para selecionar uma planilha
planilha_path = input("Por favor, insira o caminho da planilha (ex: caminho/para/seu/arquivo.xlsx): ")

# Passo 2: Carregar a planilha e obter dados
workbook = openpyxl.load_workbook(planilha_path)
sheet = workbook.active
time.sleep(2)
# Passo 3: Iterar sobre as linhas da planilha
for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    nome, telefone, status = row

    # Clicar na imagem 1
    pyautogui.click(pyautogui.locateOnScreen('imagem1.png'))
    logging.info(f'Clicou na imagem 1 para a linha {row_index}')

    # Inserir o dado da coluna nome
    pyautogui.write(nome)
    logging.info(f'Inseriu o nome {nome} para a linha {row_index}')

    # Apertar tecla tab
    pyautogui.press('tab')
    logging.info(f'Pressionou a tecla Tab para a linha {row_index}')

    # Inserir o dado da coluna telefone (convertendo para string)
    pyautogui.write(str(telefone))
    logging.info(f'Inseriu o telefone {telefone} para a linha {row_index}')

    # Clicar na imagem 2
    pyautogui.click(pyautogui.locateOnScreen('imagem2.png'))
    logging.info(f'Clicou na imagem 2 para a linha {row_index}')
    time.sleep(3)
    # Aguardar até que uma das duas imagens apareça na tela (timeout de 10 segundos)
    if wait_for_images(['imagem3.png', 'imagem4.png'], timeout=10):
        if pyautogui.locateOnScreen('imagem3.png'):
            # Escrever "adicionado" na coluna status
            status = "adicionado"
            time.sleep(2)
            logging.info(f'Escreveu "adicionado" para a linha {row_index}')
        elif pyautogui.locateOnScreen('imagem4.png'):
            # Clicar na imagem 5
            pyautogui.click(pyautogui.locateOnScreen('imagem5.png'))
            logging.info(f'Clicou na imagem 5 para a linha {row_index}')
            # Escrever "repetido" na planilha
            status = "repetido"
            logging.info(f'Escreveu "repetido" para a linha {row_index}')
    else:
        logging.warning(f'Nenhuma das imagens foi encontrada para a linha {row_index}')

    # Atualizar o valor na planilha
    sheet.cell(row=row_index, column=3, value=status)
    logging.info(f'Atualizou a planilha com status {status} para a linha {row_index}')

# Salvar a planilha
workbook.save(planilha_path)
logging.info('Salvou a planilha')

# Fechar o logger
logging.shutdown()
